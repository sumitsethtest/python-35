import urllib2, base64
import json
from pprint import pprint
import xlsxwriter
import time
from openpyxl import Workbook
from openpyxl import load_workbook
import os
os.system('cd C:\pyjira')


criteria = []
worksheet = None
wb = None

def getCriteria():
	global criteria
	global worksheet
	global wb
	
	wb = load_workbook("C:\pyjira\JiraDetails.xlsx" , read_only=False)
	ws = wb.get_sheet_by_name("Criteria")
	worksheet = wb.get_sheet_by_name("Results")
	for row in ws['A']:
		criteria.append(row.value)
	
def writeHeaders():
	global worksheet
	
	for row in worksheet['A1:G1000']:
		for cell in row:
			cell.value = None
	
	worksheet.cell(row=1, column=1).value = "PROJECT"
	worksheet.cell(row=1, column=2).value = "SUMMARY"
	worksheet.cell(row=1, column=3).value = "ASSIGNEE"
	worksheet.cell(row=1, column=4).value = "REPORTER"
	worksheet.cell(row=1, column=5).value = "STATUS"
	worksheet.cell(row=1, column=6).value = "CREATED"
	worksheet.cell(row=1, column=7).value = "UPDATED"


def getandWriteJiraData():
	global criteria
	global worksheet
	global wb
	
	origurl='https://jira/rest/api/2/search?jql='	
	row = 2
	col = 1	
	for criterion in criteria:
		url = origurl + criterion
		print url
		fullurl = urllib2.quote(url,safe="%/:=&?~#+!$,;'@()*[]")
		request = urllib2.Request(fullurl)
		base64string = base64.b64encode('%s:%s' % ('', ''))
		request.add_header("Authorization", "Basic %s" % base64string)
		result = urllib2.urlopen(request)
		cont=result.read()
		decoded = json.loads(cont)
		totalcnt=decoded['total']
		print totalcnt
		for i in range(totalcnt):
			worksheet.cell(row=row, column=col).value = decoded['issues'][i]['key']
			worksheet.cell(row=row, column=col + 1).value = decoded['issues'][i]['fields']['summary']
			worksheet.cell(row=row, column=col + 2).value = decoded['issues'][i]['fields']['assignee']['displayName']
			worksheet.cell(row=row, column=col + 3).value = decoded['issues'][i]['fields']['reporter']['displayName']
			worksheet.cell(row=row, column=col + 4).value = decoded['issues'][i]['fields']['status']['statusCategory']['name']
			worksheet.cell(row=row, column=col + 5).value = decoded['issues'][i]['fields']['created'][:10]
			worksheet.cell(row=row, column=col + 6).value = decoded['issues'][i]['fields']['updated'][:10]
			row += 1
		#row += 1

	wb.save("C:\pyjira\JiraDetails.xlsx")


if __name__ == "__main__":
	getCriteria()
	writeHeaders()
	getandWriteJiraData()
	#wb.save("C:\pyjira\JiraDetails.xlsx")
