import sys
import re
import urllib2
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import logging
import time
import traceback
import requests
from time import sleep
from selenium import webdriver

match_text=re.compile('Domain Name|Domain|licensed|ftpquota',re.IGNORECASE)
headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/50.0.2661.102 Safari/537.36'}

test_log = "test" + time.strftime("%Y%m%d-%H%M%S") + ".log"
logging.basicConfig(filename=test_log,level=logging.INFO)

results=[]
driver = webdriver.Ie("C:\Python27\IEDriverServer.exe")

logging.info("INITIATING WEB SCRAPING")
html_string = 'BLANK'
#match_text=re.compile('Advisory',re.IGNORECASE)
wb = load_workbook(filename = 'ws_input.xlsx')


def web_scrape():	
	global wb
	global results
	sheet_ranges = wb['Sheet1']
	for row in range(1,217):
		for col in range(1,2):
			print(sheet_ranges.cell(row=row,column=col).value)
			logging.info(sheet_ranges.cell(row=row,column=col).value)
			if (sheet_ranges.cell(row=row,column=col).value):
                                results=[]
				html_string = parse_html(sheet_ranges.cell(row=row,column=col).value)
				logging.info(html_string)
				cellval = html_string
				nextcol = col+1
				sheet_ranges.cell(row=row,column=nextcol).value = cellval
##				third_col = nextcol+1
##				if "Securities and|" in cellval:
##                                        sheet_ranges.cell(row=row,column=third_col).value = ""
##                                elif "" in cellval:
##                                        sheet_ranges.cell(row=row,column=third_col).value = ""
				
	wb.save(filename = 'ws_input.xlsx')


def parse_html(url):
        global results
        global driver
	page_url = url
	try:
                sleep(3)
                result = requests.get(page_url, headers=headers)
                if (result.status_code > 400):
                        #return "HTTP ISSUE - STATUS CODE" + " " + str(result.status_code)
                        result.raise_for_status()
                else:
                        soup = BeautifulSoup(result.content,'html.parser')
                        find_string_all = soup.find_all(text=re.compile('xx',re.IGNORECASE))
                        logging.info(find_string_all)
                        print(find_string_all)
                        if not (find_string_all):                                
                                driver.get(page_url)
                                html=driver.page_source
                                soup = BeautifulSoup(html,'html.parser')
                                find_string_all = soup.find_all(text=re.compile('xx',re.IGNORECASE))
                                results.extend(find_string_all)                                
                                y=driver.find_elements_by_css_selector('frame')
                                for x in y:
                                        driver.switch_to_default_content()
                                        logging.info(x)
                                        driver.switch_to_frame(x)
                                        html=driver.page_source
                                        soup = BeautifulSoup(html)
                                        find_string_all = soup.find_all(text=re.compile('xx',re.IGNORECASE))
                                        #for th in find_string_all:
                                        results.extend(find_string_all)
                                #driver.close()
                                #driver.quit()
                                return re.sub('<[^<]+?>', '', str(results))
                        return re.sub('<[^<]+?>', '', str(find_string_all))
        except requests.exceptions.ConnectionError:                
                return "WEBSITE ISSUE"
        except requests.exceptions.RequestException as e:
                #logging.error(result.raise_for_status())
                logging.error(e)
                if (result.status_code == 503):
                        return "WEB PAGE BLOCKED BY COMPANY: STATUS CODE" + " " + str(result.status_code)
                else:
                        return "HTTP ISSUE - BAD REQUEST - STATUS CODE" + " " + str(result.status_code)
	except:
		logging.error(sys.exc_info()[1])		
		return(sys.exc_info()[1])

try:
	web_scrape()
	driver.close()
	driver.quit()
except:
	logging.error(sys.exc_info()[1])
	wb.save(filename = 'ws_input.xlsx')
	driver.close()
	driver.quit()
	
	
